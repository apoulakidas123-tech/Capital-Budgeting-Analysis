import pandas as pd
import numpy as np
import numpy_financial as npf
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parent
DATA_PATH = PROJECT_ROOT / r"data\Tables_A.1_A.2_Teekay_13_25_year_capital_budgeting_plan.xls"
RESULTS_DIR = PROJECT_ROOT / "results"
RESULTS_IRR_XLSX = RESULTS_DIR / "teekay_irr_scenarios.xlsx"

SHEET_NAME = "Teekay_13"
NET_CASH_FLOW_LABEL = "Net Cash Flow"


def load_teekay_13_cash_flows(
    excel_path: str = DATA_PATH,
    sheet_name: str = SHEET_NAME,
    row_label: str = NET_CASH_FLOW_LABEL,
) -> np.ndarray:
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

    row_idx = df.index[df.iloc[:, 0] == row_label]
    if len(row_idx) == 0:
        raise ValueError(f"Row label '{row_label}' not found in sheet '{sheet_name}'.")

    row_idx = row_idx[0]
    cash_flows = df.iloc[row_idx, 1:].astype(float).values
    cash_flows = cash_flows[~np.isnan(cash_flows)]
    return cash_flows


def run_irr_scenarios_teekay13(
    cash_flows: np.ndarray,
    cost_of_capital_scenarios=(0.07, 0.05, 0.02),
) -> pd.DataFrame:
    irr_value = npf.irr(cash_flows)

    rows = []
    for rate in cost_of_capital_scenarios:
        if rate == 0.07:
            scenario_name = "7%"
        elif rate == 0.05:
            scenario_name = "5%"
        elif rate == 0.02:
            scenario_name = "2%"
        else:
            scenario_name = f"{rate:.2%}"

        npv_value = npf.npv(rate, cash_flows)

        rows.append(
            {
                "Scenario": scenario_name,
                "CostOfCapital": rate,
                "IRR": irr_value,
                "NPV_at_CostOfCapital": npv_value,
            }
        )

    df = pd.DataFrame(rows, columns=["Scenario", "CostOfCapital", "IRR", "NPV_at_CostOfCapital"])

    print("-----------------------------------------------------------------")
    print("  IRR SCENARIOS - TEEKAY 13-YEAR PROJECT")
    print("-----------------------------------------------------------------")
    print(df.to_string(index=False))
    print("-----------------------------------------------------------------")

    return df


if __name__ == "__main__":
    cash_flows_13 = load_teekay_13_cash_flows()
    irr_df = run_irr_scenarios_teekay13(cash_flows_13)

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    # Also export a nicely formatted Excel file to avoid header/data overlap
    with pd.ExcelWriter(RESULTS_IRR_XLSX, engine="openpyxl") as writer:
        irr_df.to_excel(writer, index=False, sheet_name="Scenarios")
        workbook = writer.book
        sheet = workbook["Scenarios"]

        # Adjust column widths based on content length
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            sheet.column_dimensions[column_letter].width = max_length + 2

        # Center and wrap all data cells
        from openpyxl.styles import Alignment, Font

        data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = data_alignment

        # Make header row bold and centered
        header_alignment = Alignment(horizontal="center", vertical="center")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = header_alignment

    print(f"\nIRR scenarios saved to '{RESULTS_IRR_XLSX.name}' in the 'results' folder.")

